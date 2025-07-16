import platform
import subprocess
import ipaddress
import socket
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from collections import defaultdict

class NetworkScanner:
    def __init__(self, cidr="/24"):
        self.local_ip = self.get_local_ip()
        self.network = ipaddress.ip_network(self.local_ip + cidr, strict=False)
        self.active_hosts = []

    def get_local_ip(self):
        """Récupère l'adresse IP locale."""
        hostname = socket.gethostname()
        return socket.gethostbyname(hostname)

    def ping(self, ip):
        """Teste le ping d'une IP."""
        param = '-n' if platform.system().lower() == 'windows' else '-c'
        try:
            subprocess.check_output(['ping', param, '1', ip], stderr=subprocess.DEVNULL)
            return ip, True
        except subprocess.CalledProcessError:
            return ip, False

    def get_hostname(self, ip):
        """Tente de récupérer le hostname."""
        try:
            return socket.gethostbyaddr(ip)[0]
        except:
            return "Inconnu"

    def scan(self):
        """Scanne le réseau et remplit self.active_hosts"""
        print(f"📡 Adresse IP locale détectée : {self.local_ip}")
        print(f"🔍 Scan du réseau {self.network}...\n")

        with ThreadPoolExecutor(max_workers=100) as executor:
            futures = [executor.submit(self.ping, str(ip)) for ip in self.network.hosts()]
            for future in futures:
                ip, is_up = future.result()
                if is_up:
                    hostname = self.get_hostname(ip)
                    self.active_hosts.append((hostname, ip))
                    print(f"✅ {hostname} ({ip}) est actif")

        print(f"\n🎯 Total des hôtes actifs : {len(self.active_hosts)}")

    def detect_ip_duplicates(self):
        """Détecte les IP en double."""
        ip_map = defaultdict(list)
        for hostname, ip in self.active_hosts:
            ip_map[ip].append(hostname)

        duplicates = {ip: hostnames for ip, hostnames in ip_map.items() if len(hostnames) > 1}

        if duplicates:
            print("\n⚠️ IPs en double détectées :")
            for ip, hostnames in duplicates.items():
                print(f"🔁 {ip} utilisé par : {', '.join(hostnames)}")
        else:
            print("\n✅ Aucune IP en double détectée.")

        return duplicates

    def export_to_excel(self, filename="active_hosts.xlsx"):
        """Exporte les résultats vers un fichier Excel avec doublons en rouge."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Hôtes Actifs"

        ws.append(["Nom de machine", "Adresse IP", "Date / Heure"])
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Détecter les doublons
        duplicates = self.detect_ip_duplicates()
        duplicated_ips = set(duplicates.keys())

        for hostname, ip in self.active_hosts:
            row = [hostname, ip, now]
            ws.append(row)
            if ip in duplicated_ips:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        wb.save(filename)
        print(f"\n📁 Résultats exportés dans le fichier : {filename}")

# Lancement du script
if __name__ == "__main__":
    scanner = NetworkScanner()
    scanner.scan()
    scanner.export_to_excel()
